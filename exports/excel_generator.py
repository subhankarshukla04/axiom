"""
Excel DCF Model Generator
Generates multi-sheet .xlsx workbook with:
  Sheet 1 — DCF Model (10-year projection + terminal value)
  Sheet 2 — Sensitivity Analysis (WACC vs Terminal Growth 2D table)
  Sheet 3 — Comparable Companies
  Sheet 4 — Assumptions Audit (with source labels)
  Sheet 5 — LBO Model (if LBO data available)
"""

import io
import logging
from datetime import datetime
from typing import Optional

logger = logging.getLogger(__name__)


def _get_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, Alignment, PatternFill, Border, Side, numbers
        )
        return openpyxl
    except ImportError:
        raise ImportError(
            'openpyxl is required for Excel export. '
            'Install it with: pip install openpyxl'
        )


# ── Color Palette ──────────────────────────────────────────────────────────────
HEADER_FILL = 'FF1A2036'      # Dark navy
SUBHEADER_FILL = 'FF2D4270'   # Medium navy
HIGHLIGHT_FILL = 'FF003366'   # Accent blue
GREEN_FILL = 'FF00695C'
RED_FILL = 'FFB71C1C'
YELLOW_FILL = 'FFF9A825'
LIGHT_FILL = 'FFE8EEF4'
WHITE = 'FFFFFFFF'

HEADER_FONT_COLOR = 'FFFFFFFF'
BODY_FONT = 'Calibri'
MONO_FONT = 'Courier New'


def _style_header(cell, text: str, bold: bool = True, size: int = 11):
    from openpyxl.styles import Font, Alignment, PatternFill
    cell.value = text
    cell.font = Font(name=BODY_FONT, bold=bold, color=HEADER_FONT_COLOR, size=size)
    cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')


def _style_label(cell, text: str):
    from openpyxl.styles import Font, PatternFill, Alignment
    cell.value = text
    cell.font = Font(name=BODY_FONT, bold=True, size=10)
    cell.fill = PatternFill(start_color=LIGHT_FILL, end_color=LIGHT_FILL, fill_type='solid')
    cell.alignment = Alignment(horizontal='left')


def _style_number(cell, value, fmt: str = '#,##0.0'):
    from openpyxl.styles import Alignment, Font
    cell.value = value
    cell.number_format = fmt
    cell.font = Font(name=MONO_FONT, size=10)
    cell.alignment = Alignment(horizontal='right')


def _style_signal(cell, value: float, current_price: float):
    from openpyxl.styles import Font, PatternFill, Alignment
    if current_price and current_price > 0:
        upside = (value - current_price) / current_price
        if upside >= 0.20:
            fill_color = '00C853'
        elif upside >= 0.10:
            fill_color = '64DD17'
        elif upside >= -0.10:
            fill_color = 'FFD600'
        elif upside >= -0.20:
            fill_color = 'FF6D00'
        else:
            fill_color = 'D50000'
    else:
        fill_color = 'FFFFFF'

    cell.value = round(value, 2) if value else None
    cell.number_format = '$#,##0.00'
    cell.font = Font(name=MONO_FONT, size=9, bold=True)
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    cell.alignment = Alignment(horizontal='center')


def _build_dcf_sheet(ws, company: dict, valuation: dict):
    """Build Sheet 1: DCF Model."""
    from openpyxl.styles import Font, Alignment, PatternFill

    ws.column_dimensions['A'].width = 36
    for col in 'BCDEFGHIJKLM':
        ws.column_dimensions[col].width = 12

    # ── Title ─────────────────────────────────────────────────────────────────
    ws['A1'] = f"AXIOM — DCF Model: {company.get('name', 'Unknown')} ({company.get('ticker', '')})"
    ws['A1'].font = Font(name=BODY_FONT, bold=True, size=14)
    ws['A2'] = f"Valuation Date: {datetime.utcnow().strftime('%Y-%m-%d')}  |  For Internal Use Only"
    ws['A2'].font = Font(name=BODY_FONT, italic=True, size=10, color='666666')

    # ── Assumptions Section ───────────────────────────────────────────────────
    ws['A4'] = 'ASSUMPTIONS'
    ws['A4'].font = Font(name=BODY_FONT, bold=True, size=12)

    assumptions = [
        ('Revenue Growth Y1', valuation.get('growth_rate_y1', 0), 'User Input', '0.0%'),
        ('Revenue Growth Y2', valuation.get('growth_rate_y2', 0), 'User Input', '0.0%'),
        ('Revenue Growth Y3', valuation.get('growth_rate_y3', 0), 'User Input', '0.0%'),
        ('Terminal Growth Rate', valuation.get('terminal_growth', 0.025), 'User Input', '0.00%'),
        ('WACC', valuation.get('wacc', 0.09), 'Computed (CAPM)', '0.00%'),
        ('Tax Rate', valuation.get('tax_rate', 0.21), 'User Input', '0.0%'),
        ('Beta', valuation.get('beta', 1.0), 'yfinance 5yr regression', '0.00'),
        ('Risk-Free Rate', valuation.get('risk_free_rate', 0.044), 'FRED: DGS10', '0.00%'),
        ('Market Risk Premium', valuation.get('market_risk_premium', 0.055), 'User Input', '0.00%'),
        ('CapEx (% Revenue)', valuation.get('capex_pct', 0.05), 'SEC EDGAR XBRL / yfinance', '0.0%'),
    ]

    for i, (label, value, source, fmt) in enumerate(assumptions, start=5):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = Font(name=BODY_FONT, size=10)
        ws[f'B{i}'] = value
        ws[f'B{i}'].number_format = fmt
        ws[f'B{i}'].font = Font(name=MONO_FONT, size=10)
        ws[f'C{i}'] = f'[{source}]'
        ws[f'C{i}'].font = Font(name=BODY_FONT, size=9, color='888888', italic=True)

    # ── Projections Header ────────────────────────────────────────────────────
    header_row = 16
    ws[f'A{header_row}'] = 'INCOME STATEMENT PROJECTIONS ($M)'
    ws[f'A{header_row}'].font = Font(name=BODY_FONT, bold=True, size=12)

    year_cols = list('BCDEFGHIJK')
    years = list(range(1, 11))
    ws[f'A{header_row+1}'] = ''
    for j, (col, yr) in enumerate(zip(year_cols, years)):
        _style_header(ws[f'{col}{header_row+1}'], f'Year {yr}', size=10)

    # ── Projection Rows ───────────────────────────────────────────────────────
    proj = valuation.get('projections', {})
    revenue_proj = proj.get('revenue', [0] * 10)
    fcf_proj = proj.get('fcf', [0] * 10)
    ebitda_proj = proj.get('ebitda', [0] * 10)

    metrics = [
        ('Revenue ($M)', revenue_proj, '#,##0.0'),
        ('EBITDA ($M)', ebitda_proj, '#,##0.0'),
        ('Free Cash Flow ($M)', fcf_proj, '#,##0.0'),
    ]

    for m_idx, (metric_name, values, fmt) in enumerate(metrics, start=header_row+2):
        ws[f'A{m_idx}'] = metric_name
        ws[f'A{m_idx}'].font = Font(name=BODY_FONT, size=10, bold=True if m_idx % 3 == 0 else False)
        for j, (col, val) in enumerate(zip(year_cols, values)):
            v = val / 1e6 if val and val > 1e6 else val  # Convert to $M if in raw dollars
            _style_number(ws[f'{col}{m_idx}'], v, fmt)

    # ── DCF Summary ───────────────────────────────────────────────────────────
    sum_row = header_row + len(metrics) + 3
    ws[f'A{sum_row}'] = 'DCF VALUATION SUMMARY'
    ws[f'A{sum_row}'].font = Font(name=BODY_FONT, bold=True, size=12)

    dcf_summary = [
        ('Sum of PV FCFs ($M)', valuation.get('pv_fcfs_total', 0), '#,##0.0'),
        ('Terminal Value ($M)', valuation.get('terminal_value', 0), '#,##0.0'),
        ('PV of Terminal Value ($M)', valuation.get('pv_terminal', 0), '#,##0.0'),
        ('Enterprise Value ($M)', valuation.get('enterprise_value', 0), '#,##0.0'),
        ('Less: Net Debt ($M)', valuation.get('net_debt', 0), '#,##0.0'),
        ('Equity Value ($M)', valuation.get('equity_value', 0), '#,##0.0'),
        ('Shares Outstanding (M)', valuation.get('shares_outstanding', 0) / 1e6 if valuation.get('shares_outstanding') else 0, '#,##0.0'),
        ('Implied Share Price', valuation.get('fair_value', valuation.get('dcf_value', 0)), '$#,##0.00'),
        ('Current Market Price', valuation.get('current_price', 0), '$#,##0.00'),
        ('Upside / (Downside)',
         (valuation.get('fair_value', 0) - valuation.get('current_price', 0)) / max(valuation.get('current_price', 1), 1),
         '0.0%'),
    ]

    for i, (label, value, fmt) in enumerate(dcf_summary, start=sum_row+1):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = Font(name=BODY_FONT, size=10, bold=(label == 'Implied Share Price'))
        ws[f'B{i}'] = value
        ws[f'B{i}'].number_format = fmt
        ws[f'B{i}'].font = Font(name=MONO_FONT, size=10, bold=(label == 'Implied Share Price'))


def _build_sensitivity_sheet(ws, sensitivity_data: dict):
    """Build Sheet 2: WACC vs Terminal Growth sensitivity table."""
    from openpyxl.styles import Font

    ws.column_dimensions['A'].width = 20
    for col in 'BCDEFGH':
        ws.column_dimensions[col].width = 14

    ws['A1'] = 'SENSITIVITY ANALYSIS — Implied Share Price'
    ws['A1'].font = Font(name=BODY_FONT, bold=True, size=13)
    ws['A2'] = 'Rows: Terminal Growth Rate  |  Columns: WACC  |  Green = Buy, Yellow = Hold, Red = Sell'
    ws['A2'].font = Font(name=BODY_FONT, size=10, italic=True, color='666666')

    if not sensitivity_data or 'cells' not in sensitivity_data:
        ws['A4'] = 'Sensitivity data not available — run valuation first'
        return

    x_vals = sensitivity_data['x_axis']['formatted']
    y_vals = sensitivity_data['y_axis']['formatted']
    cells = sensitivity_data['cells']
    current_price = sensitivity_data.get('current_price', 0)

    # Headers
    ws['A4'] = f"{sensitivity_data['y_axis']['label']} \\ {sensitivity_data['x_axis']['label']}"
    ws['A4'].font = Font(name=BODY_FONT, bold=True, size=10)
    for j, label in enumerate(x_vals):
        col_letter = chr(ord('B') + j)
        _style_header(ws[f'{col_letter}4'], label, size=9)

    for i, (y_label, row) in enumerate(zip(y_vals, cells), start=5):
        ws[f'A{i}'] = y_label
        ws[f'A{i}'].font = Font(name=BODY_FONT, bold=True, size=10)
        for j, cell_data in enumerate(row):
            col_letter = chr(ord('B') + j)
            val = cell_data.get('value')
            if val is not None:
                _style_signal(ws[f'{col_letter}{i}'], val, current_price)
            else:
                ws[f'{col_letter}{i}'] = 'N/A'


def _build_assumptions_audit_sheet(ws, company: dict, valuation: dict):
    """Build Sheet 4: Assumptions audit with data source labels."""
    from openpyxl.styles import Font

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20

    _style_header(ws['A1'], 'Assumption', size=11)
    _style_header(ws['B1'], 'Value', size=11)
    _style_header(ws['C1'], 'Source', size=11)
    _style_header(ws['D1'], 'Fetched At', size=11)

    rows = [
        ('Risk-Free Rate (10Y Treasury)', valuation.get('risk_free_rate', ''), 'FRED: DGS10', ''),
        ('Market Risk Premium', valuation.get('market_risk_premium', ''), 'User Input', ''),
        ('Beta (5yr vs SPY)', valuation.get('beta', ''), 'yfinance regression', ''),
        ('WACC (computed)', valuation.get('wacc', ''), 'Computed (CAPM)', ''),
        ('Revenue (LTM)', company.get('revenue', ''), 'SEC EDGAR XBRL / yfinance', ''),
        ('EBITDA (LTM)', company.get('ebitda', ''), 'SEC EDGAR XBRL / yfinance', ''),
        ('Shares Outstanding', valuation.get('shares_outstanding', ''), 'yfinance sharesOutstanding', ''),
        ('Net Debt', valuation.get('net_debt', ''), 'SEC EDGAR XBRL / yfinance', ''),
        ('Terminal Growth Rate', valuation.get('terminal_growth', ''), 'User Input (guardrailed)', ''),
        ('Y1 Growth Rate', valuation.get('growth_rate_y1', ''), 'User Input', ''),
        ('Y2 Growth Rate', valuation.get('growth_rate_y2', ''), 'User Input', ''),
        ('Y3 Growth Rate', valuation.get('growth_rate_y3', ''), 'User Input', ''),
        ('CapEx % Revenue', valuation.get('capex_pct', ''), 'SEC EDGAR XBRL / yfinance', ''),
        ('Tax Rate', valuation.get('tax_rate', ''), 'User Input', ''),
    ]

    for i, (assumption, value, source, fetched) in enumerate(rows, start=2):
        ws[f'A{i}'] = assumption
        ws[f'A{i}'].font = Font(name=BODY_FONT, size=10)
        ws[f'B{i}'] = value
        ws[f'B{i}'].font = Font(name=MONO_FONT, size=10)
        ws[f'C{i}'] = source
        ws[f'C{i}'].font = Font(name=BODY_FONT, size=10, color='2196F3')
        ws[f'D{i}'] = fetched or datetime.utcnow().strftime('%Y-%m-%d')
        ws[f'D{i}'].font = Font(name=BODY_FONT, size=9, color='888888')


def generate_excel(
    company: dict,
    valuation: dict,
    sensitivity_data: Optional[dict] = None,
    lbo_data: Optional[dict] = None,
    comps_data: Optional[list] = None,
) -> bytes:
    """
    Generate Excel DCF workbook as bytes.

    Args:
        company: dict with name, ticker, sector, current_price, etc.
        valuation: dict with all valuation outputs (dcf_value, wacc, projections, etc.)
        sensitivity_data: Optional sensitivity table from sensitivity.py
        lbo_data: Optional LBO result dict
        comps_data: Optional list of comparable companies

    Returns:
        Excel file as bytes (stream to HTTP response).
    """
    openpyxl = _get_openpyxl()
    wb = openpyxl.Workbook()

    # ── Sheet 1: DCF Model ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'DCF Model'
    _build_dcf_sheet(ws1, company, valuation)

    # ── Sheet 2: Sensitivity Analysis ─────────────────────────────────────────
    ws2 = wb.create_sheet('Sensitivity Analysis')
    _build_sensitivity_sheet(ws2, sensitivity_data or {})

    # ── Sheet 3: Comparable Companies ─────────────────────────────────────────
    ws3 = wb.create_sheet('Comparable Companies')
    ws3['A1'] = 'Comparable Companies Analysis'
    from openpyxl.styles import Font as OxFont
    ws3['A1'].font = OxFont(name=BODY_FONT, bold=True, size=13)
    if comps_data:
        headers = ['Company', 'Ticker', 'EV/EBITDA', 'P/E', 'Revenue ($M)', 'Source']
        for j, h in enumerate(headers, start=1):
            _style_header(ws3.cell(row=3, column=j), h)
        for i, comp in enumerate(comps_data, start=4):
            ws3.cell(row=i, column=1).value = comp.get('name', '')
            ws3.cell(row=i, column=2).value = comp.get('ticker', '')
            ws3.cell(row=i, column=3).value = comp.get('ev_ebitda')
            ws3.cell(row=i, column=4).value = comp.get('pe')
            ws3.cell(row=i, column=5).value = comp.get('revenue', 0)
            ws3.cell(row=i, column=6).value = comp.get('source', 'yfinance')
    else:
        ws3['A3'] = 'No comparable companies data available — add peers to generate this sheet'

    # ── Sheet 4: Assumptions Audit ────────────────────────────────────────────
    ws4 = wb.create_sheet('Assumptions Audit')
    _build_assumptions_audit_sheet(ws4, company, valuation)

    # ── Sheet 5: LBO Model (if available) ─────────────────────────────────────
    if lbo_data:
        ws5 = wb.create_sheet('LBO Model')
        ws5['A1'] = f"LBO Analysis — {company.get('name', '')} ({company.get('ticker', '')})"
        ws5['A1'].font = OxFont(name=BODY_FONT, bold=True, size=13)
        ws5['A3'] = 'Entry Structure'
        ws5['A3'].font = OxFont(name=BODY_FONT, bold=True, size=11)
        lbo_metrics = [
            ('Entry EV ($M)', lbo_data.get('entry_ev')),
            ('Entry Equity ($M)', lbo_data.get('entry_equity')),
            ('Exit EV ($M)', lbo_data.get('exit_ev')),
            ('Exit Equity ($M)', lbo_data.get('exit_equity')),
            ('IRR', f"{lbo_data.get('irr', 0):.1f}%"),
            ('MOIC', f"{lbo_data.get('moic', 0):.2f}x"),
            ('IRR Signal', lbo_data.get('irr_signal', '')),
        ]
        for i, (label, value) in enumerate(lbo_metrics, start=4):
            ws5[f'A{i}'] = label
            ws5[f'A{i}'].font = OxFont(name=BODY_FONT, size=10)
            ws5[f'B{i}'] = value
            ws5[f'B{i}'].font = OxFont(name=MONO_FONT, size=10)

        # Debt paydown schedule
        ws5['A12'] = 'DEBT PAYDOWN SCHEDULE'
        ws5['A12'].font = OxFont(name=BODY_FONT, bold=True, size=11)
        schedule = lbo_data.get('debt_paydown_schedule', [])
        if schedule:
            headers = ['Year', 'Revenue', 'EBITDA', 'Interest', 'FCF', 'Paydown', 'Remaining Debt']
            for j, h in enumerate(headers, start=1):
                _style_header(ws5.cell(row=13, column=j), h, size=9)
            for i, row in enumerate(schedule, start=14):
                for j, key in enumerate(['year', 'revenue', 'ebitda', 'interest', 'fcf', 'paydown', 'remaining_debt'], start=1):
                    ws5.cell(row=i, column=j).value = row.get(key)

    # ── Footer ────────────────────────────────────────────────────────────────
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = True

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
